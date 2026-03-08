"""
GET /api/export — export filtered jobs as XLSX.
Accepts the same filter params as /api/jobs.
"""

import io
from datetime import datetime, timedelta, date as date_type
from typing import List, Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from database import Job, get_session

router = APIRouter(prefix="/api/export", tags=["export"])

JOB_TYPE_EN = {
    "软件": "Software Eng",
    "数据工程": "Data Engineering",
    "数据分析": "Data Analytics",
    "数据科学": "Data Science",
    "AI": "ML / AI",
    "云运维": "Cloud & DevOps",
    "安全": "Cybersecurity",
    "产品": "Product",
    "商业": "Business Analysis",
    "定量": "Quantitative",
    "其他": "Other",
    # backward compat for old records
    "数据": "Data",
}

HEADERS = [
    "Date", "Company", "Type", "Job Type", "Job Title",
    "Location", "Salary", "Visa Status", "Company Size",
    "Skills", "Education", "Link",
]

COL_WIDTHS = [12, 22, 14, 10, 50, 20, 24, 28, 22, 45, 18, 70]

VISA_COLORS = {
    "✅": "FFD5F5D5",  # green
    "⚠️": "FFFFF3CD",  # orange
    "🟡": "FFFFF9C3",  # yellow
    "❌": "FFFFD5D5",  # red
}


def _visa_fill(visa_status: str) -> Optional[PatternFill]:
    for emoji, color in VISA_COLORS.items():
        if emoji in (visa_status or ""):
            return PatternFill(fill_type="solid", fgColor=color)
    return None


@router.get("")
async def export_jobs(
    q: Optional[str] = None,
    visa: List[str] = Query(default=[]),
    location: List[str] = Query(default=[]),
    job_type: List[str] = Query(default=[]),
    skills: List[str] = Query(default=[]),
    days: int = Query(default=7, ge=1, le=365),
    db: AsyncSession = Depends(get_session),
):
    cutoff_dt = datetime.utcnow() - timedelta(days=days)
    cutoff_date = cutoff_dt.date()
    conditions = [
        Job.is_active == True,
        or_(
            Job.posted_date >= cutoff_date,
            and_(Job.posted_date == None, Job.scraped_at >= cutoff_dt),
        ),
    ]

    if q:
        search_term = f"%{q}%"
        conditions.append(or_(
            Job.title.ilike(search_term),
            Job.company.ilike(search_term),
            Job.location.ilike(search_term),
        ))
    if visa:
        conditions.append(or_(*[Job.visa_status.ilike(f"%{v}%") for v in visa]))
    if location:
        conditions.append(or_(*[Job.location.ilike(f"%{loc}%") for loc in location]))
    if job_type:
        conditions.append(or_(*[Job.job_type == jt for jt in job_type]))
    if skills:
        conditions.append(or_(*[Job.skills.ilike(f"%{sk}%") for sk in skills]))

    stmt = select(Job).where(and_(*conditions)).order_by(Job.posted_date.desc().nulls_last())
    result = await db.execute(stmt)
    jobs = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "GradJobsUK"

    # Header row
    header_fill = PatternFill(fill_type="solid", fgColor="FF8CDDFA")
    header_font = Font(bold=True)
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Set column widths
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    link_font = Font(color="FF175CEB", underline="single")
    for row_idx, job in enumerate(jobs, 2):
        date_str = job.posted_date.strftime("%m-%d") if job.posted_date else ""
        row_data = [
            date_str,
            job.company,
            job.project_type,
            JOB_TYPE_EN.get(job.job_type, job.job_type),
            job.title,
            job.location,
            job.salary,
            job.visa_status,
            job.company_size,
            job.skills,
            job.education,
            job.link,
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Visa column coloring
            if col_idx == 8 and value:
                fill = _visa_fill(value)
                if fill:
                    cell.fill = fill

            # Link column styling
            if col_idx == 12 and value:
                cell.font = link_font
                cell.hyperlink = value

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"GradJobsUK_{datetime.utcnow().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
