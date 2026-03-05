"""
Export job data to Excel (XLSX) format (V2)
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from config import (
    EXCEL_COLUMNS,
    COLUMN_WIDTHS,
    HEADER_COLOR,
    LINK_COLOR,
    VISA_YES_COLOR,
    VISA_NO_COLOR,
)


def export_to_excel(jobs_data, output_filename=None):
    """
    Export job data to Excel file

    Args:
        jobs_data: List of job dictionaries
        output_filename: Optional custom filename

    Returns:
        Filepath of created Excel file
    """
    if output_filename is None:
        today = datetime.now().strftime("%Y-%m-%d")
        output_filename = f"UK_jobs_{today}.xlsx"

    # Ensure output directory exists
    output_dir = "output"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    output_path = os.path.join(output_dir, output_filename)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Listings"

    # Write header
    for col_num, column_title in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title

        # Header styling
        header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
        header_font = Font(name="Arial", size=10, bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Set column widths
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Write data rows (V2 - new column structure)
    for row_num, job in enumerate(jobs_data, 2):
        # A: 更新时间 (Update date)
        ws.cell(row=row_num, column=1).value = job.get("update_date", datetime.now().strftime("%Y-%m-%d"))

        # B: 公司名称 (Company)
        ws.cell(row=row_num, column=2).value = job.get("company", "")

        # C: 项目类型 (Project type)
        ws.cell(row=row_num, column=3).value = job.get("project_type", "")

        # D: 岗位类型 (Job type)
        ws.cell(row=row_num, column=4).value = job.get("job_type", "")

        # E: 岗位名称 (Job title)
        ws.cell(row=row_num, column=5).value = job.get("title", "")

        # F: 工作地区 (Location)
        ws.cell(row=row_num, column=6).value = job.get("location", "")

        # G: 💰 薪资 (Salary) - V2 NEW
        ws.cell(row=row_num, column=7).value = job.get("salary", "未标明")

        # H: 🔑 签证/工签 (Visa Status) - V2 NEW
        visa_cell = ws.cell(row=row_num, column=8)
        visa_status = job.get("visa_status", "未说明")
        visa_cell.value = visa_status

        # Apply conditional formatting for visa column
        if "可提供工签" in visa_status:
            visa_cell.fill = PatternFill(start_color=VISA_YES_COLOR, end_color=VISA_YES_COLOR, fill_type="solid")
        elif "不提供工签" in visa_status:
            visa_cell.fill = PatternFill(start_color=VISA_NO_COLOR, end_color=VISA_NO_COLOR, fill_type="solid")

        # I: 🏢 公司规模 (Company Size) - V2 NEW
        ws.cell(row=row_num, column=9).value = job.get("company_size", "未知")

        # J: 📋 岗位关键词 (Skill Keywords) - V2 NEW
        ws.cell(row=row_num, column=10).value = job.get("skills", "/")

        # K: 学历要求 (Education requirement)
        ws.cell(row=row_num, column=11).value = job.get("education", "官网无说明")

        # L: link
        link_cell = ws.cell(row=row_num, column=12)
        link_url = job.get("link", "")
        if link_url:
            link_cell.value = link_url
            link_cell.hyperlink = link_url
            link_font = Font(name="Arial", size=10, color=LINK_COLOR, underline="single")
            link_cell.font = link_font
        else:
            link_cell.value = ""

        # Data row styling
        data_font = Font(name="Arial", size=10)
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for col_num in range(1, len(EXCEL_COLUMNS) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.alignment = data_alignment

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save workbook
    wb.save(output_path)

    return output_path


def get_output_filename(custom_name=None):
    """
    Generate output filename

    Args:
        custom_name: Optional custom filename

    Returns:
        Filename string
    """
    if custom_name:
        return custom_name

    today = datetime.now().strftime("%Y-%m-%d")
    return f"UK_jobs_{today}.xlsx"
